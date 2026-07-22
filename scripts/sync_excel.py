#!/usr/bin/env python3
"""Przepisuje oceny/komentarze/nowe piwa ze wspólnej bazy (jsonblob) do Piwa.xlsx.

Uruchamiany co godzinę przez GitHub Actions (.github/workflows/sync.yml).

Odporność na znikanie bazy (jsonblob potrafi skasować blob):
  * adres bazy trzymany jest w store-config.json (a nie na sztywno),
  * po każdym udanym odczycie zapisujemy wierną kopię do store-backup.json,
  * jeśli baza zwróci 404, odtwarzamy ją z store-backup.json pod nowym adresem
    i aktualizujemy store-config.json (system sam się leczy w ciągu godziny).

Kolumny arkusza "Ocenka": A lp, B marka, C nazwa, D %, E rodzaj, F OCENA,
G uwagi/komentarze, H link, I komentarze www, J oceny szczegółowo.

Zmienne środowiskowe do testów lokalnych:
  STORE_FILE — czytaj bazę z pliku zamiast z sieci (pomija też zapis sieciowy)
"""
import json
import os
import sys

import openpyxl

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
XLSX = os.path.join(ROOT, "Piwa.xlsx")
CONFIG = os.path.join(ROOT, "store-config.json")
BACKUP = os.path.join(ROOT, "store-backup.json")
FALLBACK_URL = "https://jsonblob.com/api/jsonBlob/019f8998-9caa-7192-84a6-70fde04098e1"

COL_LP, COL_MARKA, COL_NAZWA, COL_ABV, COL_RODZAJ = 1, 2, 3, 4, 5
COL_OCENA, COL_UWAGI, COL_LINK, COL_KOM_WWW, COL_OCENY = 6, 7, 8, 9, 10


def norm(s):
    return " ".join(str(s if s is not None else "").lower().split())


def beer_id(marka, nazwa, rodzaj):
    return norm(marka) + "|" + norm(nazwa) + "|" + norm(rodzaj)


def fmt(n):
    return f"{n:.2f}".rstrip("0").rstrip(".").replace(".", ",")


def norm_store(s):
    s = s if isinstance(s, dict) else {}
    return {"ratings": s.get("ratings") or {}, "comments": s.get("comments") or {}, "newBeers": s.get("newBeers") or []}


def read_config_url():
    try:
        with open(CONFIG, encoding="utf-8") as f:
            url = json.load(f).get("storeUrl")
            if url:
                return url
    except (OSError, ValueError):
        pass
    return FALLBACK_URL


def write_config_url(url):
    with open(CONFIG, "w", encoding="utf-8") as f:
        json.dump({"storeUrl": url}, f, ensure_ascii=False, indent=2)
        f.write("\n")


def read_backup():
    try:
        with open(BACKUP, encoding="utf-8") as f:
            return norm_store(json.load(f))
    except (OSError, ValueError):
        return norm_store({})


def write_backup(store):
    with open(BACKUP, "w", encoding="utf-8") as f:
        json.dump(norm_store(store), f, ensure_ascii=False, indent=2)
        f.write("\n")


def get_store():
    """Zwraca (store, url). Odtwarza bazę z kopii, jeśli zniknęła (404)."""
    path = os.environ.get("STORE_FILE")
    if path:
        with open(path, encoding="utf-8") as f:
            return norm_store(json.load(f)), None

    import requests
    url = read_config_url()
    r = requests.get(url, headers={"Accept": "application/json"}, timeout=30)
    if r.status_code == 404:
        print(f"Baza pod {url} zniknęła (404) — odtwarzam z store-backup.json…")
        backup = read_backup()
        cr = requests.post("https://jsonblob.com/api/jsonBlob",
                           headers={"Content-Type": "application/json", "Accept": "application/json"},
                           data=json.dumps(backup, ensure_ascii=False).encode("utf-8"), timeout=30)
        cr.raise_for_status()
        new_url = "https://jsonblob.com" + cr.headers["Location"]
        write_config_url(new_url)
        print(f"Odtworzono bazę pod nowym adresem: {new_url}")
        return backup, new_url
    r.raise_for_status()
    return norm_store(r.json()), url


def put_store(url, store):
    if os.environ.get("STORE_FILE") or not url:
        return
    import requests
    r = requests.put(url, headers={"Content-Type": "application/json"},
                     data=json.dumps(store, ensure_ascii=False).encode("utf-8"), timeout=30)
    r.raise_for_status()


def main():
    store, url = get_store()
    ratings = store["ratings"]
    comments = store["comments"]
    new_beers = store["newBeers"]

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Ocenka"]

    rows_by_id = {}
    max_lp = 0
    last_row = 1
    for r in range(2, ws.max_row + 1):
        marka = ws.cell(row=r, column=COL_MARKA).value
        nazwa = ws.cell(row=r, column=COL_NAZWA).value
        if not marka and not nazwa:
            continue
        last_row = r
        rows_by_id[beer_id(marka, nazwa, ws.cell(row=r, column=COL_RODZAJ).value)] = r
        lp = ws.cell(row=r, column=COL_LP).value
        if isinstance(lp, (int, float)):
            max_lp = max(max_lp, int(lp))

    changed = False
    synced_new = []

    for nb in new_beers:
        bid = nb.get("id") or beer_id(nb.get("marka"), nb.get("nazwa"), nb.get("rodzaj"))
        if bid in rows_by_id:
            synced_new.append(bid)
            continue
        last_row += 1
        max_lp += 1
        ws.cell(row=last_row, column=COL_LP, value=max_lp)
        ws.cell(row=last_row, column=COL_MARKA, value=nb.get("marka") or "")
        ws.cell(row=last_row, column=COL_NAZWA, value=nb.get("nazwa") or "")
        if isinstance(nb.get("abv"), (int, float)):
            ws.cell(row=last_row, column=COL_ABV, value=nb["abv"])
        ws.cell(row=last_row, column=COL_RODZAJ, value=nb.get("rodzaj") or "Inne")
        if nb.get("addedBy"):
            ws.cell(row=last_row, column=COL_UWAGI, value=f"dodane przez: {nb['addedBy']}")
        rows_by_id[bid] = last_row
        synced_new.append(bid)
        changed = True
        print(f"+ nowe piwo: {nb.get('marka')} {nb.get('nazwa')}")

    def set_cell(row, col, value):
        nonlocal changed
        cur = ws.cell(row=row, column=col).value
        if isinstance(cur, float) and isinstance(value, (int, float)):
            if abs(cur - value) < 1e-9:
                return
        elif cur == value or (cur in (None, "") and value in (None, "")):
            return
        ws.cell(row=row, column=col, value=value)
        changed = True

    for bid, user_ratings in ratings.items():
        row = rows_by_id.get(bid)
        if not row or not user_ratings:
            continue
        vals = [v for v in user_ratings.values() if isinstance(v, (int, float))]
        if not vals:
            continue
        set_cell(row, COL_OCENA, round(sum(vals) / len(vals), 2))
        set_cell(row, COL_OCENY, "; ".join(f"{name}: {fmt(v)}" for name, v in sorted(user_ratings.items())))

    for bid, clist in comments.items():
        row = rows_by_id.get(bid)
        if not row or not clist:
            continue
        set_cell(row, COL_KOM_WWW,
                 " | ".join(f"{c.get('author', '?')}: {c.get('text', '')} ({c.get('date', '')})" for c in clist))

    if changed:
        wb.save(XLSX)
        print("Zapisano Piwa.xlsx")
    else:
        print("Brak zmian w Piwa.xlsx")

    # kopia zapasowa bazy w repo (wierny mirror — z niej odtwarzamy przy awarii)
    write_backup(store)

    # usuń z bazy nowe piwa, które trafiły już do Excela (na świeżej kopii bazy)
    if synced_new and url:
        try:
            fresh, url2 = get_store()
            url = url2 or url
            before = len(fresh["newBeers"])
            fresh["newBeers"] = [
                nb for nb in fresh["newBeers"]
                if (nb.get("id") or beer_id(nb.get("marka"), nb.get("nazwa"), nb.get("rodzaj"))) not in synced_new
            ]
            if len(fresh["newBeers"]) != before:
                put_store(url, fresh)
                write_backup(fresh)
                print(f"Wyczyszczono {before - len(fresh['newBeers'])} zsynchronizowanych nowych piw z bazy")
        except Exception as e:  # noqa: BLE001 — czyszczenie jest opcjonalne
            print(f"Ostrzeżenie: nie udało się wyczyścić newBeers: {e}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
